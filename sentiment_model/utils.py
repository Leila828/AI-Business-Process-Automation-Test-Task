
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def get_color_for_sentiment(sentiment):
    """Return a color based on the sentiment."""
    if sentiment == "POSITIVE":
        return "Green"
    elif sentiment == "NEGATIVE":
        return "Red"
    elif sentiment == "NEUTRAL":
        return "Gray"
    else:
        return "White"  # Default if sentiment is undefined
def get_color_fill(sentiment):
    """
    Return an openpyxl PatternFill object for the given sentiment.
    """
    if sentiment == "POSITIVE":
        return PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
    elif sentiment == "NEGATIVE":
        return PatternFill(start_color="FF6961", end_color="FF6961", fill_type="solid")  # Light Red
    else:
        return PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light Gray

def apply_color(file_path):
    """
    Apply color coding to rows in the Excel file based on sentiment.
    """
    # Load the workbook and active sheet
    wb = load_workbook(file_path)
    sheet = wb.active

    # Identify sentiment column
    sentiment_col = None
    for col_idx, col in enumerate(sheet.iter_cols(1, sheet.max_column), start=1):
        if col[0].value == "Sentiment":
            sentiment_col = col_idx
            break

    if sentiment_col is None:
        raise ValueError("The 'Sentiment' column was not found in the output file.")

    # Apply colors to rows based on sentiment
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  # Skip header row
        sentiment = row[sentiment_col - 1].value  # Adjust for zero-based index
        fill = get_color_fill(sentiment)
        for cell in row:
            cell.fill = fill

    # Save changes to the workbook
    wb.save(file_path)